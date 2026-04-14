#!/usr/bin/env python3
import argparse
import csv
import shutil
import zipfile
from pathlib import Path
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

ALLOWED_FILE_EXTS = ('.csv', '.xlsx', '.zip')
PREFERRED_EXT_ORDER = {'.csv': 0, '.xlsx': 1, '.zip': 2}
REQUEST_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) GeoSphere-Institutions-Fetcher/1.0',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
}

AISHE_FALLBACK_XLSX_URLS = [
    'https://he.nic.in/aishereport/assets/excel/AISHE%20Final%20Report%202021-22.xlsx',
    'https://he.nic.in/aishereport/assets/excel/AISHE%20Final%20Report%202020-21.xlsx',
    'https://he.nic.in/aishereport/assets/excel/AISHE%20Final%20Report%202019-20.xlsx',
]

UDISE_YEAR_MASTER_URL = 'https://api.udiseplus.gov.in/open-services/v1.1/acad-year-master/public'
UDISE_SUMMARY_URL = 'https://api.udiseplus.gov.in/open-services/v1.1/schools-summarised-stats/public'


def parse_args():
    parser = argparse.ArgumentParser(description='Auto-fetch UDISE/AISHE files from source pages and output CSV files')
    parser.add_argument('--udise-url', required=True)
    parser.add_argument('--aishe-url', required=True)
    parser.add_argument('--out-dir', default='data_pipeline/downloads')
    parser.add_argument('--udise-out', default='udise_wb.csv')
    parser.add_argument('--aishe-out', default='aishe_wb.csv')
    parser.add_argument('--max-depth', type=int, default=2)
    parser.add_argument('--max-pages', type=int, default=20)
    return parser.parse_args()


def normalize_url(url: str) -> str:
    parsed = urlparse(url)
    return parsed._replace(fragment='').geturl()


def should_follow_link(current_host: str, target_host: str, url: str) -> bool:
    if not target_host or target_host == current_host:
        return True

    lower = url.lower()
    if any(token in lower for token in ('download', 'report', 'data', 'csv', 'excel', 'stat', 'udise', 'aishe')):
        return True

    trusted_domains = ('gov.in', 'nic.in', 's3waas.gov.in')
    return any(target_host.endswith(domain) for domain in trusted_domains)


def candidate_score(url: str, label: str):
    lower = url.lower()
    score = 0

    keywords_common = ['download', 'report', 'stat', 'data', 'excel', 'csv', 'publication']
    keywords_label = [label, 'west-bengal', 'west_bengal', 'wb', 'state']

    for keyword in keywords_common:
        if keyword in lower:
            score += 2

    for keyword in keywords_label:
        if keyword in lower:
            score += 4

    ext = Path(urlparse(url).path).suffix.lower()
    score += 10 - PREFERRED_EXT_ORDER.get(ext, 10)

    return score


def discover_file_links(start_url: str, label: str, max_depth: int, max_pages: int):
    start_url = normalize_url(start_url)
    start_host = urlparse(start_url).netloc
    start_ext = Path(urlparse(start_url).path).suffix.lower()
    if start_ext in ALLOWED_FILE_EXTS:
        return [start_url]

    to_visit = [(start_url, 0)]
    visited = set()
    file_candidates = set()

    while to_visit and len(visited) < max_pages:
        current, depth = to_visit.pop(0)
        if current in visited:
            continue
        visited.add(current)

        try:
            response = requests.get(current, timeout=20, headers=REQUEST_HEADERS, allow_redirects=True)
            response.raise_for_status()
        except Exception:
            continue

        content_type = (response.headers.get('content-type') or '').lower()
        if 'text/html' not in content_type and '<html' not in response.text.lower()[:500]:
            ext = Path(urlparse(current).path).suffix.lower()
            if ext in ALLOWED_FILE_EXTS:
                file_candidates.add(current)
            continue

        soup = BeautifulSoup(response.text, 'html.parser')

        for anchor in soup.find_all('a', href=True):
            href = anchor['href'].strip()
            if not href:
                continue

            absolute = normalize_url(urljoin(current, href))
            parsed = urlparse(absolute)
            ext = Path(parsed.path).suffix.lower()

            if ext in ALLOWED_FILE_EXTS:
                file_candidates.add(absolute)
                continue

            if depth >= max_depth:
                continue

            if not should_follow_link(start_host, parsed.netloc, absolute):
                continue

            lower = absolute.lower()
            if any(token in lower for token in [label, 'download', 'report', 'stat', 'data', 'excel', 'csv']):
                to_visit.append((absolute, depth + 1))

    return sorted(file_candidates, key=lambda url: candidate_score(url, label), reverse=True)


def download_to_temp(url: str, out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    filename = Path(urlparse(url).path).name or 'downloaded_file'
    target = out_dir / filename

    response = requests.get(url, timeout=60, stream=True, headers=REQUEST_HEADERS, allow_redirects=True)
    response.raise_for_status()

    with target.open('wb') as fh:
        for chunk in response.iter_content(chunk_size=1024 * 1024):
            if chunk:
                fh.write(chunk)

    return target


def xlsx_to_csv(xlsx_path: Path, output_csv: Path):
    workbook = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    output_csv.parent.mkdir(parents=True, exist_ok=True)

    with output_csv.open('w', newline='', encoding='utf-8') as fh:
        writer = csv.writer(fh)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(['' if cell is None else cell for cell in row])


def choose_from_zip(zip_path: Path, work_dir: Path):
    with zipfile.ZipFile(zip_path, 'r') as archive:
        members = [m for m in archive.namelist() if Path(m).suffix.lower() in ('.csv', '.xlsx')]
        if not members:
            return None

        members.sort(key=lambda m: ('.csv' not in m.lower(), len(m)))
        chosen = members[0]
        extracted = work_dir / Path(chosen).name
        with archive.open(chosen) as src, extracted.open('wb') as dst:
            shutil.copyfileobj(src, dst)
        return extracted


def materialize_csv_from_url(url: str, output_csv: Path, work_dir: Path) -> bool:
    ext = Path(urlparse(url).path).suffix.lower()

    try:
        downloaded = download_to_temp(url, work_dir)
    except Exception:
        return False

    try:
        if ext == '.csv':
            output_csv.parent.mkdir(parents=True, exist_ok=True)
            shutil.copyfile(downloaded, output_csv)
            return True

        if ext == '.xlsx':
            xlsx_to_csv(downloaded, output_csv)
            return True

        if ext == '.zip':
            extracted = choose_from_zip(downloaded, work_dir)
            if not extracted:
                return False
            extracted_ext = extracted.suffix.lower()
            if extracted_ext == '.csv':
                output_csv.parent.mkdir(parents=True, exist_ok=True)
                shutil.copyfile(extracted, output_csv)
                return True
            if extracted_ext == '.xlsx':
                xlsx_to_csv(extracted, output_csv)
                return True
            return False

        return False
    except Exception:
        return False


def fetch_one(label: str, source_url: str, out_csv: Path, out_dir: Path, max_depth: int, max_pages: int):
    candidates = discover_file_links(source_url, label, max_depth=max_depth, max_pages=max_pages)

    if not candidates:
        return False, []

    for candidate in candidates:
        if materialize_csv_from_url(candidate, out_csv, out_dir):
            return True, candidates

    return False, candidates


def fetch_aishe_fallback(out_csv: Path, out_dir: Path) -> bool:
    for url in AISHE_FALLBACK_XLSX_URLS:
        if materialize_csv_from_url(url, out_csv, out_dir):
            return True
    return False


def fetch_udise_fallback(out_csv: Path) -> bool:
    try:
        year_response = requests.get(UDISE_YEAR_MASTER_URL, timeout=20, headers=REQUEST_HEADERS)
        year_response.raise_for_status()
        year_data = year_response.json().get('data') or []
        if not year_data:
            return False

        year_ids = []
        for row in year_data:
            try:
                year_ids.append(int(row.get('yearId')))
            except Exception:
                continue
        if not year_ids:
            return False
        latest_year = max(year_ids)

        payload = {
            'yearId': latest_year,
            'regionType': 11,
            'regionCode': 19,
            'dType': 11,
            'dCode': 19,
            'categoryCode': 0,
            'managementCode': 0,
            'locationCode': 0,
            'schoolTypeCode': 0,
        }

        summary_response = requests.post(UDISE_SUMMARY_URL, json=payload, timeout=30, headers=REQUEST_HEADERS)
        summary_response.raise_for_status()
        summary_payload = summary_response.json()
        rows = summary_payload.get('data') or []
        if not rows:
            return False

        fieldnames = sorted({key for row in rows for key in row.keys()})
        if not fieldnames:
            return False

        out_csv.parent.mkdir(parents=True, exist_ok=True)
        with out_csv.open('w', newline='', encoding='utf-8') as fh:
            writer = csv.DictWriter(fh, fieldnames=fieldnames)
            writer.writeheader()
            for row in rows:
                writer.writerow(row)

        return True
    except Exception:
        return False


def main():
    args = parse_args()
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    udise_out = out_dir / args.udise_out
    aishe_out = out_dir / args.aishe_out

    udise_ok, udise_candidates = fetch_one('udise', args.udise_url, udise_out, out_dir, args.max_depth, args.max_pages)
    aishe_ok, aishe_candidates = fetch_one('aishe', args.aishe_url, aishe_out, out_dir, args.max_depth, args.max_pages)

    if not aishe_ok:
        aishe_ok = fetch_aishe_fallback(aishe_out, out_dir)

    if not udise_ok:
        udise_ok = fetch_udise_fallback(udise_out)

    print(f'UDISE_OK={str(udise_ok).lower()}')
    print(f'AISHE_OK={str(aishe_ok).lower()}')
    print(f'UDISE_OUT={udise_out}')
    print(f'AISHE_OUT={aishe_out}')

    if not udise_ok:
        print('UDISE_CANDIDATES=' + '|'.join(udise_candidates[:10]))
    if not aishe_ok:
        print('AISHE_CANDIDATES=' + '|'.join(aishe_candidates[:10]))

    if not (udise_ok and aishe_ok):
        raise SystemExit(1)


if __name__ == '__main__':
    main()
